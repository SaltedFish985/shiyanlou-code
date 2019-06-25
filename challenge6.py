# -*- coding: utf-8 -*-

from openpyxl import load_workbook 
from openpyxl import Workbook 
import datetime 

def combine():
    the_xlsx = load_workbook('/home/shiyanlou/Code/courses.xlsx')
    the_stutents = the_xlsx.get_sheet_by_name('students')
    the_time = the_xlsx.get_sheet_by_name('time')
    the_combine = the_xlsx.create_sheet('combine')
    for column in the_stutents["A:C"]:
        for cell in column:
            the_combine[cell.column + str(cell.row)] = cell.value
    the_column = "B"
    for row1 in range(2,486):            
        for row2 in range(2,486):
            if the_combine[the_column + str(row1)].value == the_time[the_column + str(row2)].value:
                the_combine["D" + str(row1)].value = the_time["C" + str(row2)].value
    the_combine["D1"] = "学习时间"
    the_xlsx.save('/home/shiyanlou/Code/courses.xlsx')

def split():
    the_year = []
    the_xlsx = load_workbook('/home/shiyanlou/Code/courses.xlsx')
    the_combine = the_xlsx.get_sheet_by_name('combine')
    the_column = "A"
    for row in range(2, 486):
        the_year.append(the_combine[the_column + str(row)].value.year)
    years = list(set(the_year))
    years.sort()
    for each_year in years:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = str(each_year)
        that_row = 2
        for row in range(2, 486):
            if each_year == the_combine[the_column + str(row)].value.year:                
                ws1["A" + str(that_row)].value = the_combine[row][0].value
                ws1["B" + str(that_row)].value = the_combine[row][1].value
                ws1["C" + str(that_row)].value = the_combine[row][2].value
                ws1["D" + str(that_row)].value = the_combine[row][3].value
                that_row += 1
        ws1["A1"] = "创建时间"
        ws1["B1"] = "课程名称"
        ws1["C1"] = "学习人数"
        ws1["D1"] = "学习时间"
        wb.save('/home/shiyanlou/Code/{}.xlsx'.format(str(each_year)))


if __name__ == '__main__':
    combine()
    split()
